from tools.compute_subjects_list import compute_subjects_list
import pathlib
import os
from evaluation.rmse import RMSE
from evaluation.drmse import dRMSE
from evaluation.cg_ega.cg_ega import CG_EGA
import numpy as np
import openpyxl
from evaluation.r import r
from evaluation.fit import fit
from evaluation.time_lag import time_lag
import pickle
import misc
import pandas as pd


class ResultsAnalyzer():
    """
        ResultsAnalyzer object that instiate results objects for a given population and compute the overall
        results.
        Parameters:
            - model_name: name of the model (e.g., "ELM")
            - ph: prediciton horizon (e.g., 30)
            - freq: sampling frequency (e.g., 5)
    """

    def __init__(self, model_name, ph, freq=misc.freq):
        self.ph = ph
        self.model_name = model_name
        self.freq = freq

    def analyze(self, dataset="all", subject="all"):
        """
            Compute the overall results of the given population
            :param dataset: name of the dataset, supports "all"
            :param subject: name of the subject, supports "all"
            :return: dict of mean results accros population, dict of std results accros population
        """
        datasets_subjects = compute_subjects_list(dataset, subject)

        res_list = []
        for dataset, subject in datasets_subjects:
            # load results
            file_name = "ph" + str(self.ph) + "_" + dataset + subject + "_" + self.model_name + "_.res"
            file_path = os.path.join(os.path.dirname(__file__), "..", "results", "ph" + str(self.ph), self.model_name,
                                     file_name)
            res = Results(self.model_name, self.ph, dataset, subject, self.freq, file=file_path)

            # get and store results
            res_list.append(res.get_results())

        df = pd.DataFrame(res_list, columns=list(res_list[0].keys()))

        return dict(df.mean()), dict(df.std())


class Results():
    """
        Results object that contains all the different metrics for on subject.
        Parameters:
            - model_name: name of the model (e.g., "ELM")
            - ph: prediction horizon in minutes (e.g., 30)
            - dataset: name of the dataset (e.g, "Ohio")
            - subject: name of the subject from the dataset (e.g., "591")
            - freq: sampling frequency in minutes (e.g., 5)
            - results: if not None, results are computed from the given array
            - file: if not None, results are computed from the given file, handles "auto"
    """

    def __init__(self, model_name, ph, dataset, subject, freq, results=None, file=None):
        if file is not None:
            if file == "auto":
                file_name = "ph" + str(ph) + "_" + dataset + subject + "_" + model_name + "_.res"
                self.load(os.path.join("results", "ph" + str(ph), model_name, file_name))
            else:
                self.load(file)

        elif results is not None:
            # reshape the results so that day have the shape (split, 2, day, val)
            self.results = [np.rollaxis(np.array(res), 2, 0) for res in results]

        self.model = model_name
        self.freq = freq
        self.ph = ph
        self.dataset = dataset
        self.subject = subject

    def get_results(self):
        """
            Compute the overall results, averaged for every split
            :return: dict with the following keys: RMSE, dRMSE, r, fit, TL, AP, BE, EP
        """
        # RMSE
        rmse = np.mean([RMSE(*res) for res in self.results])

        # dRMSE
        drmse = np.mean([dRMSE(*res, self.freq) for res in self.results])

        # CG-EGA
        cg_ega = np.mean([CG_EGA(*res, self.freq).reduced() for res in self.results], axis=0)

        r_metric = np.mean([r(*res) for res in self.results])
        fit_metric = np.mean([fit(*res) for res in self.results])
        time_lag_metric = np.mean([time_lag(*res) for res in self.results])

        return {
            "RMSE": rmse,
            "dRMSE": drmse,
            "r": r_metric,
            "fit": fit_metric,
            "TL": time_lag_metric,
            "AP": cg_ega[0],
            "BE": cg_ega[1],
            "EP": cg_ega[2],
        }

    def plot(self, split=0, day=0):
        """
            Plot the results of a given day from a given split
            :param split: split number
            :param day: day number
            :return: /
        """
        y_true = self.results[split][0]
        y_pred = self.results[split][1]

        CG_EGA(y_true, y_pred, self.freq).plot(day=day)

    def to_excel(self, params, n_splits, file_name="results.xlsx"):
        """
            Save the results into an excel file
            :param params:
            :param n_splits:
            :param file_name:
            :return:
        """
        file = os.path.join(misc.path, "results", file_name)

        results = self.get_results()
        misc_params = {"ph": self.ph, "dataset": self.dataset, "subject": self.subject, "n_splits": n_splits}

        data = {**misc_params, **params, **results}

        # if file not exist, create it with appropriate header
        if not pathlib.Path(file).is_file():
            wb = openpyxl.Workbook()
            wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        else:
            wb = openpyxl.load_workbook(file)

        if not self.model in wb.sheetnames:
            wb.create_sheet(self.model)
            ws = wb[self.model]
            ws.append(list(data.keys()))
        else:
            ws = wb[self.model]

        ws.append(list(data.values()))
        wb.save(file)

    def save(self, suffix=""):
        """
            Save results object to file
            :param suffix: suffix to append to the file name
            :return: /
        """
        file_name = "ph" + str(self.ph) + "_" + self.dataset + self.subject + "_" + self.model + "_" + suffix + ".res"
        dir = os.path.join(os.path.dirname(__file__), "..", "results", "ph" + str(self.ph), self.model)
        if not os.path.exists(dir): os.makedirs(dir)
        file = os.path.join(dir, file_name)

        with open(file, "wb") as the_file:
            pickle.dump(self.results, the_file)

    def load(self, file):
        """
            load Results object from file
            :param file: file
            :return: /
        """
        with open(file, "rb") as the_file:
            self.results = pickle.load(the_file)

    def save_day(self, split, day):
        """
            Save on particular day (ground_truth, predictions, AP, BE, EP)
            :param split: split number
            :param day: day number
            :return: /
        """
        y_true = self.results[split][0, day, :].reshape(1, -1)
        y_pred = self.results[split][1, day, :].reshape(1, -1)

        df = CG_EGA(y_true, y_pred, self.freq).per_sample()
        ap = df[df["CG_EGA"] == "AP"]
        be = df[df["CG_EGA"] == "BE"]
        ep = df[df["CG_EGA"] == "EP"]

        dir = "s" + str(split) + "_d" + str(day)
        path = os.path.join(os.path.dirname(__file__), "..", "results", self.model, dir)
        if not os.path.exists(path): os.makedirs(path)

        df.loc[:, ["time", "y_true"]].to_csv(os.path.join(path, "y_true"), sep=" ", index=False)
        df.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "y_pred"), sep=" ", index=False)
        ap.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "time_ap"), sep=" ", index=False)
        be.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "time_be"), sep=" ", index=False)
        ep.loc[:, ["time", "y_pred"]].to_csv(os.path.join(path, "time_ep"), sep=" ", index=False)
        ap.loc[:, ["y_true", "y_pred"]].to_csv(os.path.join(path, "p_ega_ap"), sep=" ", index=False)
        be.loc[:, ["y_true", "y_pred"]].to_csv(os.path.join(path, "p_ega_be"), sep=" ", index=False)
        ep.loc[:, ["y_true", "y_pred"]].to_csv(os.path.join(path, "p_ega_ep"), sep=" ", index=False)
        ap.loc[:, ["dy_true", "dy_pred"]].to_csv(os.path.join(path, "r_ega_ap"), sep=" ", index=False)
        be.loc[:, ["dy_true", "dy_pred"]].to_csv(os.path.join(path, "r_ega_be"), sep=" ", index=False)
        ep.loc[:, ["dy_true", "dy_pred"]].to_csv(os.path.join(path, "r_ega_ep"), sep=" ", index=False)
